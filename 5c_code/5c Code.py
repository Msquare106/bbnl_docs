import pandas as pd
from openpyxl import load_workbook
import warnings
import time
warnings.simplefilter('ignore')


start = time.time()
# def load_file(file_name:str, row_skip:int = 0, foot_skip:int = 0):
#     file_directory ='F:/bbnl 5c/'
#     data = pd.read_csv(file_directory+file_name, skiprows=row_skip, skipfooter=foot_skip)
#     print("Loaded File: ", file_name)
#     return data

def load_file(file_name:str, row_skip:int = 0, foot_skip:int = 0):
    file_directory ='F:/bbnl 5c/'
    data = pd.read_csv(file_directory+file_name, skiprows=row_skip, skipfooter=foot_skip, encoding='latin1')
    print("Loaded File: ", file_name)
    return data


five = load_file('5c.csv', 2)
amc = load_file('lms.csv', 3, 3)
ava = load_file('prev day ava.csv', 2, 2)
it = load_file('itpc.csv')
bb = load_file('bbnw.csv')
prev_day_1 = load_file('prev day 1.csv', 3, 3)
prev_day_2 = load_file('prev day 2.csv', 3, 3)
org = load_file('original.csv', 3, 3)
udy = load_file('udyami.csv', 3, 3)

five['Concat'] = five['State'] + five['District'] + five['Block'] + five['Panchayat'] + five['GP Location'] + five['GP Location Code']

amc['Concat'] = amc['State'] + amc['District'] + amc['Block'] + amc['GP'] + amc['Location Name'] + amc['Loc Code']

# For GPs in AMC as per LMS
amc1 = amc

ava['PANCHAYAT'] = ava['PANCHAYAT'].fillna('--')
ava['Concat'] = ava['STATE'] + ava['DISTRICT'] + ava['BLOCK'] + ava['PANCHAYAT'] + ava['ONT LOCATION NAME'] + ava['LOCATION CODE']

merged = pd.merge(five, amc[['Concat', 'AMC2 Start Date']], on='Concat', how='left')
merged = pd.merge(merged, ava[['Concat', 'ONT AVAILABILITY']], on='Concat', how='left')
merged = merged.drop(['APP Status', 'Physical Status', 'SPV', 'Commission Date', 'GP Added Date', 'Owner', 'Concat'], axis=1)

# Updating AMC date for PH-2 PB & BH
# merged.loc[(merged['State'].isin(['PUNJAB', 'BIHAR'])) & ((merged['AMC2 Start Date'].isna()) | (merged['AMC2 Start Date'] == '--')) & (merged['Phase'] == 'PHASE-2'), 'AMC2 Start Date'] = '01-01-2015'
# Updating AMC date for A&N
merged.loc[(merged['State'] == 'ANDAMAN AND NICOBAR ISLANDS') & ((merged['AMC2 Start Date'].isna()) | (merged['AMC2 Start Date'] == '--')), 'AMC2 Start Date'] = '01-01-2015'


def fiber_fault_count(row):
    return 5 - sum([row['TT ID Fiber-BBNL<100m'] == "--", row['TT ID Fiber-BBNL>100m And <500m'] == "--",
                    row['TT ID Fiber-BBNL>500m'] == "--", row['TT ID Leased Fiber'] == "--",
                    row['TT ID Fiber-BSNL-Lossy'] == "--"])


def no_fiber_fault(row):
    if row['Fiber Fault Count'] == 0:
        return "Yes"
    else:
        return "No"


merged['Fiber Fault Count'] = merged.apply(fiber_fault_count, axis=1)
merged_1 = merged[merged['Phase'].isin(['PHASE-1', 'PHASE 1', 'Phase-1', 'Phase 1'])]
merged_2 = merged[merged['Phase'].isin(['PHASE-2', 'PHASE 2', 'Phase-2', 'Phase 2'])]

merged_1_inamc = merged_1[~(merged_1['AMC2 Start Date'].isna()) & (merged_1['AMC2 Start Date'] != '--')]
merged_1_notinamc = merged_1[(merged_1['AMC2 Start Date'].isna()) | (merged_1['AMC2 Start Date'] == '--')]

merged_2_inamc = merged_2[~(merged_2['AMC2 Start Date'].isna()) & (merged_2['AMC2 Start Date'] != '--')]
merged_2_notinamc = merged_2[(merged_2['AMC2 Start Date'].isna()) | (merged_2['AMC2 Start Date'] == '--')]


# 1 - Creating total GPs table
final_1 = merged_1['State'].value_counts().reset_index().rename(columns={'count': 'GPs visible in NMS'}).sort_values(by='State')
final_2 = merged_2['State'].value_counts().reset_index().rename(columns={'count': 'GPs visible in NMS'}).sort_values(by='State')

# (older version of pandas 1.5.3) final_1 = merged_1['State'].value_counts().reset_index().rename(columns={'index': 'State', 'State': 'GPs visible in NMS'}).sort_values(by='State')
# (older version of pandas 1.5.3) final_2 = merged_2['State'].value_counts().reset_index().rename(columns={'index': 'State', 'State': 'GPs visible in NMS'}).sort_values(by='State')

# Adding total count for total gps table
final_1 = pd.concat([final_1, pd.DataFrame([{'State': 'Total', 'GPs visible in NMS': final_1['GPs visible in NMS'].sum()}])], ignore_index=True)
final_2 = pd.concat([final_2, pd.DataFrame([{'State': 'Total', 'GPs visible in NMS': final_2['GPs visible in NMS'].sum()}])], ignore_index=True)

# (older version of pandas 1.5.3) final_1 = final_1.append({'State': 'Total', 'GPs visible in NMS': final_1['GPs visible in NMS'].sum()}, ignore_index=True)
# (older version of pandas 1.5.3) final_2 = final_2.append({'State': 'Total', 'GPs visible in NMS': final_2['GPs visible in NMS'].sum()}, ignore_index=True)

# 2 - Creating up count table
up_gps_1 = pd.pivot_table(merged_1, index='State', values='GP Status', aggfunc=lambda x: sum(x == 'UP'), margins=True,margins_name='Total').reset_index()
up_gps_2 = pd.pivot_table(merged_2, index='State', values='GP Status', aggfunc=lambda x: sum(x == 'UP'), margins=True,margins_name='Total').reset_index()
# Creating Un_Up count table
unup_gps_1 = pd.pivot_table(merged_1, index='State', values='GP Status',aggfunc=lambda x: sum(x == 'UNKNOWN_PREVIOUSLY_UP'), margins=True,margins_name='Total').reset_index()
unup_gps_2 = pd.pivot_table(merged_2, index='State', values='GP Status',aggfunc=lambda x: sum(x == 'UNKNOWN_PREVIOUSLY_UP'), margins=True,margins_name='Total').reset_index()
# Adding up & un_up to main table
final_1['Up+Unk_Up Gps'] = up_gps_1['GP Status'] + unup_gps_1['GP Status']
final_2['Up+Unk_Up Gps'] = up_gps_2['GP Status'] + unup_gps_2['GP Status']

# 3 - Creating GP Down-No AMC-  (No Fiber Fault,Fiber Faults)
noamc_down_gps_1 = merged_1_notinamc[(merged_1_notinamc['GP Status'] == 'DOWN') | (merged_1_notinamc['GP Status'] == 'UNKNOWN_PREVIOUSLY_DOWN')]
noamc_down_gps_2 = merged_2_notinamc[(merged_2_notinamc['GP Status'] == 'DOWN') | (merged_2_notinamc['GP Status'] == 'UNKNOWN_PREVIOUSLY_DOWN')]


def count_dash(x):
    return (x != '--').sum()


def count_zero(x):
    return (x == 0).sum()


def count_not_zero(x):
    return (x != 0).sum()


categories = ['Fiber Fault Count', 'TT ID Fiber-BBNL<100m', 'TT ID Fiber-BBNL>100m And <500m', 'TT ID Fiber-BBNL>500m','TT ID Leased Fiber','TT ID Fiber-BSNL-Lossy']

aggfunc = {'Fiber Fault Count': count_zero, 'TT ID Fiber-BBNL<100m': count_dash,'TT ID Fiber-BBNL>100m And <500m': count_dash,'TT ID Fiber-BBNL>500m': count_dash, 'TT ID Leased Fiber': count_dash, 'TT ID Fiber-BSNL-Lossy': count_dash}

noamc_down_nofib_1 = pd.pivot_table(noamc_down_gps_1, index='State', values=categories, aggfunc=aggfunc, margins=True,margins_name='Total').reset_index()
noamc_down_nofib_2 = pd.pivot_table(noamc_down_gps_2, index='State', values=categories, aggfunc=aggfunc, margins=True,margins_name='Total').reset_index()
# Renaming No Fiber Fault - No AMC Date
noamc_down_nofib_1.rename(columns={'Fiber Fault Count': 'Not in AMC - No Fiber Fault'}, inplace=True)
noamc_down_nofib_2.rename(columns={'Fiber Fault Count': 'Not in AMC - No Fiber Fault'}, inplace=True)
# Adding to main table
final_1 = pd.merge(final_1, noamc_down_nofib_1, on='State', how='left')
final_2 = pd.merge(final_2, noamc_down_nofib_2, on='State', how='left')

# 4 - GPs in AMC as per LMS
amc1 = amc1.drop('Initial Phase', axis=1)
amc1 = pd.merge(amc1, five[['Concat', 'Phase']], on='Concat', how='left')

# Updating AMC date for PH-2 PB & BH
# amc1.loc[(amc1['State'].isin(['PUNJAB', 'BIHAR'])) & ((amc1['AMC2 Start Date'].isna()) | (amc1['AMC2 Start Date'] == '--')) & (amc1['Phase'] == 'PHASE-2'), 'AMC2 Start Date'] = '01-01-2015'
# Updating AMC date for A&N
amc1.loc[(amc1['State'] == 'ANDAMAN AND NICOBAR ISLANDS') & ((amc1['AMC2 Start Date'].isna()) | (amc1['AMC2 Start Date'] == '--')), 'AMC2 Start Date'] = '01-01-2015'

amc1 = amc1[amc1['AMC2 Start Date'] != '--']

amc1_1 = amc1[(amc1['Phase'].isin(['PHASE-1', 'PHASE 1', 'Phase-1', 'Phase 1'])) | (amc1['Phase'].isna())]
amc1_2 = amc1[(amc1['Phase'].isin(['PHASE-2', 'PHASE 2', 'Phase-2', 'Phase 2'])) | (amc1['Phase'].isna())]

amc1_1 = amc1_1['State'].value_counts().reset_index().rename(columns={'count': 'AMC Started as per LMS'}).sort_values(by='State')
amc1_2 = amc1_2['State'].value_counts().reset_index().rename(columns={'count': 'AMC Started as per LMS'}).sort_values(by='State')

# (older version of pandas 1.5.3) amc1_1 = amc1_1['State'].value_counts().reset_index().rename(columns={'index': 'State', 'State': 'AMC Started as per LMS'}).sort_values(by='State')
# (older version of pandas 1.5.3) amc1_2 = amc1_2['State'].value_counts().reset_index().rename(columns={'index': 'State', 'State': 'AMC Started as per LMS'}).sort_values(by='State')

# Adding total count
amc1_1 = pd.concat([amc1_1, pd.DataFrame([{'State': 'Total', 'AMC Started as per LMS': amc1_1['AMC Started as per LMS'].sum()}])], ignore_index=True)
amc1_2 = pd.concat([amc1_2, pd.DataFrame([{'State': 'Total', 'AMC Started as per LMS': amc1_2['AMC Started as per LMS'].sum()}])], ignore_index=True)

# (older version of pandas 1.5.3) amc1_1 = amc1_1.append({'State': 'Total', 'AMC Started as per LMS': amc1_1['AMC Started as per LMS'].sum()},ignore_index=True)
# (older version of pandas 1.5.3) amc1_2 = amc1_2.append({'State': 'Total', 'AMC Started as per LMS': amc1_2['AMC Started as per LMS'].sum()},ignore_index=True)

# Adding to main table
final_1 = pd.merge(final_1, amc1_1, on='State', how='left')
final_2 = pd.merge(final_2, amc1_2, on='State', how='left')

# 4 - Default Faults in GPs under AMC
categories1 = ['District', 'Fiber Fault Count', 'TT ID ONT Faulty', 'TT ID ONT Missing', 'TT ID CCU Faulty', 'TT ID CCU Missing',
               'TT ID GP Shifting', 'TT ID Solar Cable', 'TT ID SPV Mounting', 'TT ID PP Extension', 'TT ID Solar Panel Faulty', 'TT ID SPV Missing',
               'TT ID Earthing Issue', 'TT ID Battery Faulty', 'TT ID Battery Missing', 'TT ID Power', 'TT ID Electricity Issues', 'TT ID Custodian Issues']

aggfunc1 = {'District': 'count', 'Fiber Fault Count': count_not_zero, 'TT ID ONT Faulty': count_dash, 'TT ID ONT Missing': count_dash,
            'TT ID CCU Faulty': count_dash, 'TT ID CCU Missing': count_dash, 'TT ID GP Shifting': count_dash, 'TT ID Solar Cable': count_dash,
            'TT ID SPV Mounting': count_dash, 'TT ID PP Extension': count_dash, 'TT ID Solar Panel Faulty': count_dash, 'TT ID SPV Missing': count_dash,
            'TT ID Earthing Issue': count_dash, 'TT ID Battery Faulty': count_dash, 'TT ID Battery Missing': count_dash, 'TT ID Power': count_dash,
            'TT ID Electricity Issues': count_dash, 'TT ID Custodian Issues': count_dash}

flt_inamc_gps_1 = pd.pivot_table(merged_1_inamc, index='State', values=categories1, aggfunc=aggfunc1, margins=True,
                                 margins_name='Total').reset_index()
flt_inamc_gps_2 = pd.pivot_table(merged_2_inamc, index='State', values=categories1, aggfunc=aggfunc1, margins=True,
                                 margins_name='Total').reset_index()

# Renaming District - AMC Started as per Inventory
flt_inamc_gps_1.rename(columns={'District': 'Total GPs - AMC Started as per Inventory'}, inplace=True)
flt_inamc_gps_2.rename(columns={'District': 'Total GPs - AMC Started as per Inventory'}, inplace=True)
# Renaming Fiber Fault - In AMC
flt_inamc_gps_1.rename(columns={'Fiber Fault Count': 'In AMC - Fiber Fault'}, inplace=True)
flt_inamc_gps_2.rename(columns={'Fiber Fault Count': 'In AMC - Fiber Fault'}, inplace=True)
# Adding to main table
final_1 = pd.merge(final_1, flt_inamc_gps_1, on='State', how='left')
final_2 = pd.merge(final_2, flt_inamc_gps_2, on='State', how='left')

# 5 - GP Down - No Fiber Faults in GPs under AMC
amc_down_gps_1 = merged_1_inamc[
    (merged_1_inamc['GP Status'] == 'DOWN') | (merged_1_inamc['GP Status'] == 'UNKNOWN_PREVIOUSLY_DOWN')]
amc_down_gps_2 = merged_2_inamc[
    (merged_2_inamc['GP Status'] == 'DOWN') | (merged_2_inamc['GP Status'] == 'UNKNOWN_PREVIOUSLY_DOWN')]

amc_down_nofib_1 = pd.pivot_table(amc_down_gps_1, index='State', values='Fiber Fault Count', aggfunc=count_zero,
                                  margins=True, margins_name='Total').reset_index()
amc_down_nofib_2 = pd.pivot_table(amc_down_gps_2, index='State', values='Fiber Fault Count', aggfunc=count_zero,
                                  margins=True, margins_name='Total').reset_index()

# Renaming GP Down - No Fiber Fault - In AMC
amc_down_nofib_1.rename(columns={'Fiber Fault Count': 'In AMC - GP Down - No Fiber Fault'}, inplace=True)
amc_down_nofib_2.rename(columns={'Fiber Fault Count': 'In AMC - GP Down - No Fiber Fault'}, inplace=True)

# Adding to the main table
final_1 = pd.merge(final_1, amc_down_nofib_1, on='State', how='left')
final_2 = pd.merge(final_2, amc_down_nofib_2, on='State', how='left')



# 6 - Creating Total Mini OLTs count table
it_1 = it[it['Phase']=='Phase-1']
it_2 = it[it['Phase']=='Phase-2 CPSU']

itf_1 = it_1['state'].value_counts().reset_index()
itf_2 = it_2['state'].value_counts().reset_index()

itf_1.columns = ['State', 'Total Mini OLT Count']
itf_2.columns = ['State', 'Total Mini OLT Count']

itf_1 = pd.concat([itf_1, pd.DataFrame([{'State': 'Total', 'Total Mini OLT Count': itf_1['Total Mini OLT Count'].sum()}])], ignore_index=True)
itf_2 = pd.concat([itf_2, pd.DataFrame([{'State': 'Total', 'Total Mini OLT Count': itf_2['Total Mini OLT Count'].sum()}])], ignore_index=True)

# (older version of pandas 1.5.3) itf_1 = itf_1.append({'State': 'Total', 'Total Mini OLT Count': itf_1['Total Mini OLT Count'].sum()}, ignore_index=True)
# (older version of pandas 1.5.3) itf_2 = itf_2.append({'State': 'Total', 'Total Mini OLT Count': itf_2['Total Mini OLT Count'].sum()}, ignore_index=True)


# Adding to the main table
final_1 = pd.merge(final_1, itf_1, on='State', how='left')
final_2 = pd.merge(final_2, itf_2, on='State', how='left')


# 7 - Creating UP Mini OLT Table
itbb_1 = pd.merge(it_1, bb, left_on='OLT_IP', right_on='IP Address', how='left')
itbb_2 = pd.merge(it_2, bb, left_on='OLT_IP', right_on='IP Address', how='left')

itbb_1 = itbb_1[~(itbb_1['SSA Name'].isna()) & (itbb_1['BBNMS Reachability Status']=='UP')]
itbb_2 = itbb_2[~(itbb_2['SSA Name'].isna()) & (itbb_2['BBNMS Reachability Status']=='UP')]

itbb_1 = itbb_1['state'].value_counts().reset_index()
itbb_2 = itbb_2['state'].value_counts().reset_index()

itbb_1.columns = ['State', 'UP Mini OLT Count']
itbb_2.columns = ['State', 'UP Mini OLT Count']

itbb_1 = pd.concat([itbb_1, pd.DataFrame([{'State': 'Total', 'UP Mini OLT Count': itbb_1['UP Mini OLT Count'].sum()}])], ignore_index=True)
itbb_2 = pd.concat([itbb_2, pd.DataFrame([{'State': 'Total', 'UP Mini OLT Count': itbb_2['UP Mini OLT Count'].sum()}])], ignore_index=True)

# (older version of pandas 1.5.3) itbb_1 = itbb_1.append({'State': 'Total', 'UP Mini OLT Count': itbb_1['UP Mini OLT Count'].sum()}, ignore_index=True)
# (older version of pandas 1.5.3) itbb_2 = itbb_2.append({'State': 'Total', 'UP Mini OLT Count': itbb_2['UP Mini OLT Count'].sum()}, ignore_index=True)

# Adding to the main table
final_1 = pd.merge(final_1, itbb_1, on='State', how='left')
final_2 = pd.merge(final_2, itbb_2, on='State', how='left')


# 8 - Creating GPs up due to Mini OLT table
udy_1 = udy[(udy['NMS NAME'] == 'NOFN') & (udy['PIA'].isin(['BSNL_1', 'PGCIL', 'RAILTEL']))]
udy_2 = udy[(udy['NMS NAME'] == 'NOFN') & (~udy['PIA'].isin(['BSNL_1', 'PGCIL', 'RAILTEL']))]

org = org[org['NMS NAME'] == 'NOFN']

udy_1 = pd.merge(udy_1, org, on='LGD CODE', how = 'left')
udy_2 = pd.merge(udy_2, org, on='LGD CODE', how = 'left')

udy_1 = udy_1[udy_1['STATE NAME_y'].isna()]
udy_2 = udy_2[udy_2['STATE NAME_y'].isna()]

udy_1 = udy_1['STATE NAME_x'].value_counts().reset_index()
udy_2 = udy_2['STATE NAME_x'].value_counts().reset_index()

udy_1.columns = ['State', 'GP Down - Mini OLT Up Count']
udy_2.columns = ['State', 'GP Down - Mini OLT Up Count']

udy_1 = pd.concat([udy_1, pd.DataFrame([{'State': 'Total', 'GP Down - Mini OLT Up Count': udy_1['GP Down - Mini OLT Up Count'].sum()}])], ignore_index=True)
udy_2 = pd.concat([udy_2, pd.DataFrame([{'State': 'Total', 'GP Down - Mini OLT Up Count': udy_2['GP Down - Mini OLT Up Count'].sum()}])], ignore_index=True)

# (older version of pandas 1.5.3) udy_1 = udy_1.append({'State': 'Total', 'GP Down - Mini OLT Up Count': udy_1['GP Down - Mini OLT Up Count'].sum()}, ignore_index=True)
# (older version of pandas 1.5.3) udy_2 = udy_2.append({'State': 'Total', 'GP Down - Mini OLT Up Count': udy_2['GP Down - Mini OLT Up Count'].sum()}, ignore_index=True)

# Adding to the main table
final_1 = pd.merge(final_1, udy_1, on='State', how='left')
final_2 = pd.merge(final_2, udy_2, on='State', how='left')




# 9 - Up anytime during table
prev_day_1.rename(columns={'STATE_NAME': 'State'}, inplace=True)
prev_day_1['State'] = prev_day_1['State'].fillna('Total')

prev_day_2.rename(columns={'STATE_NAME': 'State'}, inplace=True)
prev_day_2['State'] = prev_day_2['State'].fillna('Total')

# Adding to the main table
final_1 = pd.merge(final_1, prev_day_1, on='State', how='left')
final_2 = pd.merge(final_2, prev_day_2, on='State', how='left')


# 10 - prev day 98% availability
mt95_inamc_1 = merged_1_inamc[merged_1_inamc['ONT AVAILABILITY']>= 98]
mt95_inamc_2 = merged_2_inamc[merged_2_inamc['ONT AVAILABILITY']>= 98]

mt95_gps_inamc_1 = mt95_inamc_1['State'].value_counts().reset_index().rename(columns={'count': 'GPs with >95% Ava'}).sort_values(by='State')
mt95_gps_inamc_2 = mt95_inamc_2['State'].value_counts().reset_index().rename(columns={'count': 'GPs with >95% Ava'}).sort_values(by='State')

# (older version of pandas 1.5.3) mt95_gps_inamc_1 = mt95_inamc_1['State'].value_counts().reset_index().rename(columns={'index': 'State', 'State': 'GPs with >95% Ava'}).sort_values(by='State')
# (older version of pandas 1.5.3) mt95_gps_inamc_2 = mt95_inamc_2['State'].value_counts().reset_index().rename(columns={'index': 'State', 'State': 'GPs with >95% Ava'}).sort_values(by='State')

mt95_gps_inamc_1 = pd.concat([mt95_gps_inamc_1, pd.DataFrame([{'State': 'Total', 'GPs with >95% Ava': mt95_gps_inamc_1['GPs with >95% Ava'].sum()}])], ignore_index=True)
mt95_gps_inamc_2 = pd.concat([mt95_gps_inamc_2, pd.DataFrame([{'State': 'Total', 'GPs with >95% Ava': mt95_gps_inamc_2['GPs with >95% Ava'].sum()}])], ignore_index=True)

# (older version of pandas 1.5.3) mt95_gps_inamc_1 = mt95_gps_inamc_1.append({'State': 'Total', 'GPs with >95% Ava': mt95_gps_inamc_1['GPs with >95% Ava'].sum()}, ignore_index=True)
# (older version of pandas 1.5.3) mt95_gps_inamc_2 = mt95_gps_inamc_2.append({'State': 'Total', 'GPs with >95% Ava': mt95_gps_inamc_2['GPs with >95% Ava'].sum()}, ignore_index=True)


# Adding to the main table
final_1 = pd.merge(final_1, mt95_gps_inamc_1, on='State', how='left')
final_2 = pd.merge(final_2, mt95_gps_inamc_2, on='State', how='left')

# 12 - AMC lms- inventory availalbe
final_1['AMC Started - Inventory NA'] = final_1['AMC Started as per LMS'] - final_1['Total GPs - AMC Started as per Inventory']
final_2['AMC Started - Inventory NA'] = final_2['AMC Started as per LMS'] - final_2['Total GPs - AMC Started as per Inventory']

final_1['Leased + Lossy'] = final_1['TT ID Leased Fiber'] + final_1['TT ID Fiber-BSNL-Lossy']
final_2['Leased + Lossy'] = final_2['TT ID Leased Fiber'] + final_2['TT ID Fiber-BSNL-Lossy']

final_1['ONT Faulty + Missing'] = final_1['TT ID ONT Faulty'] + final_1['TT ID ONT Missing']
final_2['ONT Faulty + Missing'] = final_2['TT ID ONT Faulty'] + final_2['TT ID ONT Missing']

final_1['CCU Faulty + Missing'] = final_1['TT ID CCU Faulty'] + final_1['TT ID CCU Missing']
final_2['CCU Faulty + Missing'] = final_2['TT ID CCU Faulty'] + final_2['TT ID CCU Missing']

final_1['SPV Faulty + Missing'] = final_1['TT ID Solar Panel Faulty'] + final_1['TT ID SPV Missing']
final_2['SPV Faulty + Missing'] = final_2['TT ID Solar Panel Faulty'] + final_2['TT ID SPV Missing']

final_1['Battery Faulty + Missing'] = final_1['TT ID Battery Faulty'] + final_1['TT ID Battery Missing']
final_2['Battery Faulty + Missing'] = final_2['TT ID Battery Faulty'] + final_2['TT ID Battery Missing']


final_1 = final_1[['State','GPs visible in NMS','Up+Unk_Up Gps','Total Mini OLT Count','UP Mini OLT Count','GP Down - Mini OLT Up Count',
                  'Prev Day','Prev 3 Days','Not in AMC - No Fiber Fault','TT ID Fiber-BBNL<100m','TT ID Fiber-BBNL>100m And <500m',
                  'TT ID Fiber-BBNL>500m', 'Leased + Lossy', 'AMC Started as per LMS', 'Total GPs - AMC Started as per Inventory', 'AMC Started - Inventory NA',
                  'GPs with >95% Ava', 'In AMC - Fiber Fault','In AMC - GP Down - No Fiber Fault','ONT Faulty + Missing','CCU Faulty + Missing',
                  'TT ID Solar Cable','TT ID SPV Mounting','SPV Faulty + Missing', 'TT ID Earthing Issue', 'TT ID Electricity Issues',
                   'TT ID Custodian Issues', 'TT ID PP Extension', 'TT ID GP Shifting', 'Battery Faulty + Missing', 'TT ID Power']]

final_2 = final_2[['State','GPs visible in NMS','Up+Unk_Up Gps','Total Mini OLT Count','UP Mini OLT Count','GP Down - Mini OLT Up Count',
                  'Prev Day','Prev 3 Days','Not in AMC - No Fiber Fault','TT ID Fiber-BBNL<100m','TT ID Fiber-BBNL>100m And <500m',
                  'TT ID Fiber-BBNL>500m', 'Leased + Lossy', 'AMC Started as per LMS', 'Total GPs - AMC Started as per Inventory', 'AMC Started - Inventory NA',
                  'GPs with >95% Ava', 'In AMC - Fiber Fault','In AMC - GP Down - No Fiber Fault','ONT Faulty + Missing','CCU Faulty + Missing',
                  'TT ID Solar Cable','TT ID SPV Mounting','SPV Faulty + Missing', 'TT ID Earthing Issue', 'TT ID Electricity Issues',
                   'TT ID Custodian Issues', 'TT ID PP Extension', 'TT ID GP Shifting', 'Battery Faulty + Missing', 'TT ID Power']]

final_1['State'] = final_1['State'].replace('Total', 'Totals')
final_2['State'] = final_2['State'].replace('Total', 'Totals')

directory ='F:/bbnl 5c/'

# Loading the existing Excel file for Phase-1
wb = load_workbook(directory+"5C REPORT revised format.xlsx")
ws = wb.active

# Iterate over each row in the DataFrame
for index, row in final_1.iterrows():
    state = row['State']
    # Find the matching row index in the Excel sheet
    for excel_row, excel_state in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True), start=2):
        if excel_state[0] == state:
            # Copy values from the DataFrame to the Excel sheet
            for col, value in enumerate(row, start=1):
                ws.cell(row=excel_row, column=col, value=value)
            break  # Stop searching for the current state

# Save the Excel file
wb.save(directory+"5C REPORT revised format.xlsx")



# Loading the existing Excel file for Phase-2
wb = load_workbook(directory+"5C REPORT revised format.xlsx")
ws = wb.active

# Iterate over each row in the DataFrame
for index, row in final_2.iterrows():
    state = row['State']
    # Find the matching row index in the Excel sheet
    for excel_row, excel_state in enumerate(ws.iter_rows(min_row=50, max_row=ws.max_row, min_col=1, max_col=1, values_only=True), start=50):
        if excel_state[0] == state:
            # Copy values from the DataFrame to the Excel sheet
            for col, value in enumerate(row, start=1):
                ws.cell(row=excel_row, column=col, value=value)
            break  # Stop searching for the current state

# Save the Excel file
wb.save(directory+"5C REPORT revised format.xlsx")
end = time.time()

print("\nTime Taken:", round(end - start, 2), "Seconds")