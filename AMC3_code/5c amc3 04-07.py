#!/usr/bin/env python
# coding: utf-8

# In[52]:


import pandas as pd
from openpyxl import load_workbook
import warnings
import time
import numpy as np
warnings.simplefilter('ignore')


start = time.time()
def load_file(file_name:str, row_skip:int = 0, foot_skip:int = 0):
    file_directory ='F:/abp 5c/'
    data = pd.read_csv(file_directory+file_name, skiprows=row_skip, skipfooter=foot_skip)
    print("Loaded File: ", file_name)
    return data

fivec = load_file('5c.csv', 2, 2)
lms = load_file('lms.csv', 3, 3)
avail = load_file('ava.csv', 3, 3)


# In[63]:


five = fivec
amc = lms
ava = avail

five['Concat'] = five['State'] + five['District'] + five['Block'] + five['Panchayat'] + five['GP Location'] + five['GP Location Code']

amc['Concat'] = amc['State'] + amc['District'] + amc['Block'] + amc['GP'] + amc['Location Name'] + amc['Loc Code']
amc3 = amc[ ~(  (amc['AMC3 Start Date'] == '--') | (amc['State'] == 'TEST')      )]


ava['PANCHAYAT NAME'] = ava['PANCHAYAT NAME'].replace('*NULL*', '*--*')
ava['Concat'] = ava['STATE'] + ava['DISTRICT'] + ava['BLOCK'] + ava['PANCHAYAT NAME'] + ava['ONT LOCATION NAME'] + ava['ONT LOCATION CODE']

five = five.drop(['State', 'District', 'Block', 'Panchayat', 'GP Location'], axis=1)

merged = pd.merge(amc3[['Concat', 'State', 'District', 'Block', 'GP', 'Location Name', 'AMC3 Start Date']], five, on='Concat', how='left')

merged = pd.merge(merged, ava[['Concat', 'ONT AVAILABILITY(%)']], on='Concat', how='left')

print(len(merged))

# merged.to_csv('E:/BBNL/template/5c daily report/5c amc 3/merged.csv')


# In[64]:


# Creating the main table
final = merged['State'].value_counts().reset_index().rename(columns={'count': 'GPs in AMC'}).sort_values(by='State')
# Adding total to the table
final = pd.concat([final, pd.DataFrame([{'State': 'Total', 'GPs in AMC': final['GPs in AMC'].sum()}])], ignore_index=True)

# # Creating the GPs in NMS table
# gps_in_nms = merged.groupby('State')['EMS Name'].apply(lambda x: x.notna().sum()).reset_index(name='GPs in NMS')
# # Adding total to the table
# gps_in_nms = pd.concat([gps_in_nms, pd.DataFrame([{'State': 'Total', 'GPs in NMS': gps_in_nms['GPs in NMS'].sum()}])], ignore_index=True)

# #Adding to the main table
# final = pd.merge(final, gps_in_nms, on='State', how='left')


# In[65]:


# Creating the status count table
status_table = pd.pivot_table(merged, index='State', columns='GP Status', values = 'EMS Name',
    aggfunc='count', fill_value=0, margins=True, margins_name='Total')
# Using pd.crosstab() (simpler and preferred for counts):
# status_table = pd.crosstab(merged['State'], merged['GP Status'], margins=True, margins_name='Total')
# status_table = status_table.drop(['Total'], axis=1)

# Adding to the main table
final = pd.merge(final, status_table, on='State', how='left')


# In[66]:


# Creating total ticket count table
merged['Total'] = merged['Total'].fillna(0).astype(int)
ticket_count = pd.pivot_table(merged, index='State', values='Total', aggfunc='sum', fill_value=0).reset_index()
# Renaming the Total column
ticket_count.rename(columns={'Total': 'Total Ticket Count'}, inplace=True)
# Adding total to the table
ticket_count = pd.concat([ticket_count, pd.DataFrame([{'State': 'Total', 'Total Ticket Count': ticket_count['Total Ticket Count'].sum()}])], ignore_index=True)

# Adding to the main table
final = pd.merge(final, ticket_count, on='State', how='left')
print(final)


# In[67]:


def fiber_fault_count(row):
    cols = [
        'TT ID Fiber-BBNL<100m',
        'TT ID Fiber-BBNL>100m And <500m',
        'TT ID Fiber-BBNL>500m',
        'TT ID Leased Fiber',
        'TT ID Fiber-BSNL-Lossy'
    ]
    invalid_count = row[cols].apply(lambda x: pd.isna(x) or x == "--").sum()
    return len(cols) - invalid_count

merged['Fiber Fault Count'] = merged.apply(fiber_fault_count, axis=1)

fiber_count = pd.pivot_table(merged, index='State', values='Fiber Fault Count', aggfunc='sum', fill_value=0).reset_index()
   
# Adding total to the table
fiber_count = pd.concat([fiber_count, pd.DataFrame([{'State': 'Total', 'Fiber Fault Count': fiber_count['Fiber Fault Count'].sum()}])], ignore_index=True)

# Adding to the main table
final = pd.merge(final, fiber_count, on='State', how='left')


# In[68]:


def count_valid_faults(series):
    return series.apply(lambda x: x != '--' and not pd.isna(x)).sum()

categories = [
    'TT ID ONT Faulty', 'TT ID ONT Missing', 'TT ID CCU Faulty', 'TT ID CCU Missing',
    'TT ID GP Shifting', 'TT ID Solar Cable', 'TT ID SPV Mounting', 'TT ID PP Extension',
    'TT ID Solar Panel Faulty', 'TT ID SPV Missing', 'TT ID Earthing Issue',
    'TT ID Battery Faulty', 'TT ID Battery Missing', 'TT ID Power',
    'TT ID Electricity Issues', 'TT ID Custodian Issues']

# Automatically assign the same function to all categories
aggfunc = {col: count_valid_faults for col in categories}

# Generate the pivot table
faults = pd.pivot_table(
    merged,
    index='State',
    values=categories,
    aggfunc=aggfunc,
    margins=True,
    margins_name='Total'
).reset_index()


# In[69]:


# Combining few fault categories
faults['ONT Faulty + Missing'] = faults['TT ID ONT Faulty'] + faults['TT ID ONT Missing']
faults['CCU Faulty + Missing'] = faults['TT ID CCU Faulty'] + faults['TT ID CCU Missing']
faults['SPV Faulty + Missing'] = faults['TT ID Solar Panel Faulty'] + faults['TT ID SPV Missing']
faults['Battery Faulty + Missing'] = faults['TT ID Battery Faulty'] + faults['TT ID Battery Missing']

faults = faults.drop(['TT ID ONT Faulty', 'TT ID ONT Missing', 'TT ID CCU Faulty', 'TT ID CCU Missing', 'TT ID Solar Panel Faulty',
                      'TT ID SPV Missing', 'TT ID Battery Faulty', 'TT ID Battery Missing'], axis=1)

# Adding to the main table
final = pd.merge(final, faults, on='State', how='left')


# In[70]:


# Creating gp count with ava>=98% table
avamt = merged[merged['ONT AVAILABILITY(%)'] >= 98]
avamt = avamt.groupby('State').size().reset_index(name='Count >=98%')

# Adding to the main table
final = pd.merge(final, avamt, on='State', how='left')


# In[71]:


# Creating % of '98% UP GPs' from the HOTO GPs
# final['% of 98% UP GPs'] = (final['Count >=98%'] / final['GPs in AMC']) * 100
final['% of 98% UP GPs'] = np.where(
    final['GPs in AMC'] == 0,
    '',  # keep cell empty
    round((final['Count >=98%'] / final['GPs in AMC']) * 100, 2)
)

final = final[['State', 'GPs in AMC', 'Total', 'UP', 'UNKNOWN_PREVIOUSLY_UP', 'UNKNOWN_PREVIOUSLY_DOWN', 'DOWN',
        'Count >=98%', '% of 98% UP GPs', 'Total Ticket Count', 'Fiber Fault Count', 'TT ID Custodian Issues', 'TT ID Earthing Issue', 
        'TT ID Electricity Issues', 'TT ID GP Shifting', 'TT ID PP Extension', 'TT ID Power', 'TT ID SPV Mounting',
        'TT ID Solar Cable', 'ONT Faulty + Missing', 'CCU Faulty + Missing', 'SPV Faulty + Missing',
       'Battery Faulty + Missing']]

# # Rearranging the columns as per the template
# final = final[['State', 'GPs in AMC', 'Total', 'UP', 'UNKNOWN_PREVIOUSLY_UP', 'UNKNOWN_PREVIOUSLY_DOWN', 'DOWN',
#         'Count >=98%', 'Total Ticket Count', 'Fiber Fault Count', 'TT ID Custodian Issues', 'TT ID Earthing Issue', 
#         'TT ID Electricity Issues', 'TT ID GP Shifting', 'TT ID PP Extension', 'TT ID Power', 'TT ID SPV Mounting',
#         'TT ID Solar Cable', 'ONT Faulty + Missing', 'CCU Faulty + Missing', 'SPV Faulty + Missing',
#        'Battery Faulty + Missing']]


# In[72]:


from openpyxl import load_workbook

directory = 'F:/abp 5c/'
file_path = directory + "5C amc3 template.xlsx"

# Load workbook and worksheet
wb = load_workbook(file_path)
ws = wb.active

# Step 1: Build a mapping: State name -> Excel row number
state_row_map = {}
for row in range(5, ws.max_row + 1):  # Adjust row start if needed
    state = ws.cell(row=row, column=2).value  # Column B (index 2)
    if state:
        state_row_map[state] = row

# Step 2: Write values from DataFrame, skipping 'State', starting from Column C (index 3)
for _, df_row in final.iterrows():
    state = df_row['State']
    excel_row = state_row_map.get(state)

    if excel_row:
        for col_offset, value in enumerate(df_row.drop('State').values):  # Skip 'State' column
            ws.cell(row=excel_row, column=3 + col_offset).value = value  # Start from column C

# Step 3: Save the workbook
wb.save(file_path)

