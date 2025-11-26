# driver.find_element(By.ID,"user").send_keys("root")
# driver.find_element(By.ID,"password").send_keys("Au79c0Rn3r$")
from openpyxl import Workbook, load_workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains 
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.alert import Alert
from bs4 import BeautifulSoup
import warnings
import requests
import time
import os
import glob

opt = webdriver.EdgeOptions()
opt.add_experimental_option('detach', True)
opt.add_argument("--ignore-certificate-errors")
opt.set_capability("acceptInsecureCerts", True)
driver = webdriver.Edge(options=opt)
driver.minimize_window()
print("Browser opened")

try:
    wb = load_workbook("DR_sheet -Daily.xlsx")
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active

def insert_row(ip_address, status, ws, row, two_slot_exist=False, slot_names = []):
    output_1 = " & ".join(slot_names) + ": "
    output_2 = []
    cell = ""
    if two_slot_exist:
        if status[0] == "" and status[1] == "":
            output_2.append(CellRichText(get_nic_ports("", "N/A")))
        else:
            output_2.append(CellRichText(get_nic_ports(status[0], "1"), get_nic_ports(status[1], "2")))
        if status[2] == "" and status[3] == "":
            output_2.append(CellRichText(get_nic_ports("", "N/A")))
        else:
            output_2.append(CellRichText(get_nic_ports(status[2], "1"), get_nic_ports(status[3], "2")))
        cell = CellRichText(*(list(CellRichText(get_nic_ports("", output_1)))
                              + list(output_2[0]) 
                              + list(CellRichText(get_nic_ports("", " & "))) 
                              + list(output_2[1])))
    else:
        if status[0] == "" and status[1] == "":
            cell = "N/A"
        else:
            cell = CellRichText(get_nic_ports(status[0], "1"), get_nic_ports(status[1], "2"))
    ws[f'H{row}'].value = cell
    # print(cell)
    # ws.append([ip_address, colored_cell])

    print("Updated Workbook with", ip_address, "\n")


# def insert_row_integrated(ip_address, status, ws, row, slot_names=[]):
#     output_1 = " & ".join(slot_names) + ": "
#     cell = ""

#     if status[0] == "" and status[1] == "":
#         cell = ""   # leave empty if both blank
#     else:
#         parts = []
#         if status[0] != "":
#             parts.append(get_nic_ports(status[0], "1"))
#         if status[1] != "":
#             parts.append(get_nic_ports(status[1], "2"))

#         # add slot name prefix
#         cell = CellRichText(
#             *(list(CellRichText(get_nic_ports("", output_1))) + parts)
#         )

#     ws[f'I{row}'].value = cell
#     print("Integrated Slot Updated Workbook with", ip_address, "\n")


def insert_row_integrated(ip_address, status, ws, row, slot_names=[]):
    output_1 = " & ".join(slot_names) + ": "
    parts = []

    for idx, val in enumerate(status[:4], start=1):
        if val != "":   # skip blanks, no "N/A"
            parts.append(get_nic_ports(val, str(idx)))

    if not parts:   # all four are blank
        cell = ""
    else:
        cell = CellRichText(
            *(list(CellRichText(get_nic_ports("", output_1))) + parts)
        )

    ws[f'I{row}'].value = cell
    print("Integrated Slot Updated Workbook with", ip_address, "\n")



def get_nic_ports(link_status, port):
    text_color = "FF000000"
    if str(link_status) == "Up":
        text_color = "FF00FF00"
    elif str(link_status) == "Down":
        text_color = "FFFF0000"
    return TextBlock(InlineFont(color=text_color), port+" ")

def get_info(driver, ip_address, row):
    try:
        two_slot_exist = False
        two_slot_ip = ['172.31.40.31', '172.31.40.32', '172.31.40.28', '172.31.40.29']
        if ip_address == '172.31.39.106':
            return False
        elif ip_address in two_slot_ip:
            two_slot_exist = True

        driver.get(f'https://{ip_address}')
        print("Routed to link", ip_address)

        wait=WebDriverWait(driver, 100)
        wait.until(EC.element_to_be_clickable((By.ID,"user")))

        print("Link Opened")

        time.sleep(2)
        driver.find_element(By.ID,"user").send_keys("root")
        driver.find_element(By.ID,"password").send_keys("Au79c0Rn3r$")
        driver.find_element(By.ID,"btnOK").click()

        time.sleep(2)
        print("Logged In")
        time.sleep(10)
        main_frame = driver.find_elements(By.ID, "navigationBar")
        if(len(main_frame) > 0):
            print("Server Frame Loading...")
            driver.switch_to.frame("da")

        loader = driver.find_elements(By.ID, "progressGraphic")
        if(len(loader) > 0):
            wait.until(EC.invisibility_of_element_located((By.ID, "progressGraphic")))
            print("Server Frame Loaded!")
            driver.switch_to.default_content()

        time.sleep(4)
        wait.until(EC.presence_of_element_located((By.ID, "treelist_id")))
        side_options_frame = driver.find_element(By.ID, "treelist_id")


        time.sleep(2)
        driver.switch_to.frame(side_options_frame)

        driver.execute_script("f_expand('C18')")
        print("Hardware Options Extended")

        wait.until(EC.element_to_be_clickable((By.ID, "a_Network Devices"))).click()
        print("Network Devices Frame Loaded!")

        # ---------------- NIC SLOTS (Do not change this logic) ----------------
        if not two_slot_exist:
            driver.switch_to.default_content()
            driver.switch_to.frame("snb")
            wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/table/tbody/tr/td[4]/a"))).click()
            slot_no = str(driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr/td[4]/a").get_attribute("innerHTML"))[-1]

            time.sleep(10)
            driver.switch_to.default_content()
            driver.switch_to.frame("da")

            
            nic_table = driver.find_element(By.ID, "networkdevice_ports_partition_port").find_element(By.CSS_SELECTOR, "table.container")

            link_status = [f'linkStatus_NIC.Slot.{slot_no}-1-1', f'linkStatus_NIC.Slot.{slot_no}-2-1']
            status = []

            for i in link_status:
                status.append(nic_table.find_element(By.ID, i).get_attribute("innerHTML"))
            

            insert_row(ip_address, status, ws, row)

        else:
            print("Two Slot Exist - NIC")
            status, slot_names = [], []
            for td in [4, 6]:
                driver.switch_to.default_content()
                driver.switch_to.frame("snb")
                wait.until(EC.element_to_be_clickable((By.XPATH, f'/html/body/div/table/tbody/tr/td[{td}]/a'))).click()
                slot_no = str(driver.find_element(By.XPATH, f'/html/body/div/table/tbody/tr/td[{td}]/a').get_attribute("innerHTML"))
                slot_names.append(slot_no)

                print("NIC Slot Clicked")


                time.sleep(10)
                driver.switch_to.default_content()
                driver.switch_to.frame("da")
                nic_table = driver.find_element(By.ID, "networkdevice_ports_partition_port").find_element(By.CSS_SELECTOR, "table.container")

                link_status = [f'linkStatus_NIC.Slot.{slot_no[-1]}-1-1', f'linkStatus_NIC.Slot.{slot_no[-1]}-2-1']
                for i in link_status:
                    status.append(nic_table.find_element(By.ID, i).get_attribute("innerHTML"))

            insert_row(ip_address, status, ws, row, two_slot_exist, slot_names)

        # ---------------- INTEGRATED SLOTS ----------------
        driver.switch_to.default_content()
        driver.switch_to.frame("snb")

        if not two_slot_exist:  # if condition (td[6])
            wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/table/tbody/tr/td[6]/a"))).click()
            slot_no = str(driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr/td[6]/a").get_attribute("innerHTML"))[-1]
            time.sleep(10)
        else:  # else condition (td[8])
            wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/table/tbody/tr/td[8]/a"))).click()
            slot_no = str(driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr/td[8]/a").get_attribute("innerHTML"))[-1]
            time.sleep(10)

        driver.switch_to.default_content()
        driver.switch_to.frame("da")
        integrated_table = driver.find_element(By.ID, "networkdevice_ports_partition_port").find_element(By.CSS_SELECTOR, "table.container")

        # integrated has 4 possible IDs but we only care about 2 active rows
        link_status = [
            f'linkStatus_NIC.Integrated.{slot_no}-1-1',
            f'linkStatus_NIC.Integrated.{slot_no}-2-1',
            f'linkStatus_NIC.Integrated.{slot_no}-3-1',
            f'linkStatus_NIC.Integrated.{slot_no}-4-1'
        ]

        status = []
        for i in link_status:
            try:
                status.append(integrated_table.find_element(By.ID, i).get_attribute("innerHTML"))
            except:
                status.append("")

        insert_row_integrated(ip_address, status, ws, row)

    except Exception as e:
        print("Error:", e)
        wb.save("output.xlsx")
        driver.close()


ip = [cell.value for cell in ws['F'][2:]]
for index, address in enumerate(ip):
    get_info(driver, address, int(ip.index(address))+3)

wb.save("output.xlsx")
driver.close()